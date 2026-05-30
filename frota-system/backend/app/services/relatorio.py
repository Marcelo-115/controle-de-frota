import io
import base64
from typing import Dict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _gerar_grafico_barras(maquinas: list) -> str:
    nomes = [m["nome"] for m in maquinas]
    medias = [m["media"] or 0 for m in maquinas]

    cores = []
    for v in medias:
        if v == 0:
            cores.append("#94a3b8")
        elif v < 3.0:
            cores.append("#22c55e")
        elif v <= 4.0:
            cores.append("#f59e0b")
        else:
            cores.append("#ef4444")

    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(nomes, medias, color=cores, edgecolor="white")
    ax.set_xlabel("Máquinas", fontsize=11)
    ax.set_ylabel("L/h", fontsize=11)
    ax.set_title("Consumo médio por máquina (L/h)", fontsize=13, fontweight="bold")
    ax.axhline(y=3.0, color="#3b82f6", linestyle="--", alpha=0.6, label="Limite econômico (3.0)")
    ax.axhline(y=4.0, color="#ef4444", linestyle="--", alpha=0.6, label="Limite normal (4.0)")
    ax.legend()
    ax.set_ylim(bottom=0)

    for bar, val in zip(bars, medias):
        if val > 0:
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                bar.get_height() + 0.05,
                f"{val:.2f}",
                ha="center",
                va="bottom",
                fontsize=9,
            )

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def gerar_excel(dados: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    titulo_font = Font(bold=True, size=14, color="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    kpi_font = Font(bold=True, size=12)

    ws["A1"] = "RELATÓRIO DE GESTÃO DE FROTA"
    ws["A1"].font = titulo_font
    ws["A2"] = f"Período: {dados.get('periodo', '')}"
    ws["A2"].font = Font(italic=True, size=10)

    ws["A4"] = "KPIs do Período"
    ws["A4"].font = Font(bold=True, size=11)

    kpis = [
        ("Total de Máquinas", dados.get("total_maquinas", 0)),
        ("Total de Horas", f"{dados.get('total_horas', 0):.1f} h"),
        ("Total de Litros", f"{dados.get('total_litros', 0):.1f} L"),
        ("Média Geral", f"{dados.get('media_geral', 0):.2f} L/h"),
    ]
    for i, (label, value) in enumerate(kpis, 5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value).font = kpi_font

    headers = ["Máquina", "Tipo", "Horas (h)", "Litros (L)", "Média (L/h)", "Status", "Abastecimentos"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_map = {"Econômico": "22c55e", "Normal": "f59e0b", "Alto": "ef4444", "sem_dados": "94a3b8"}

    for row_idx, maquina in enumerate(dados.get("maquinas", []), 11):
        ws.cell(row=row_idx, column=1, value=maquina.get("nome"))
        ws.cell(row=row_idx, column=2, value=maquina.get("tipo"))
        ws.cell(row=row_idx, column=3, value=maquina.get("horas")).number_format = "0.0"
        ws.cell(row=row_idx, column=4, value=maquina.get("litros")).number_format = "0.0"
        media_cell = ws.cell(row=row_idx, column=5, value=maquina.get("media"))
        media_cell.number_format = "0.00"
        status_val = maquina.get("status", "sem_dados")
        status_cell = ws.cell(row=row_idx, column=6, value=status_val)
        cor = status_map.get(status_val, "94a3b8")
        status_cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        status_cell.font = Font(color="FFFFFF", bold=True)
        status_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=maquina.get("qtd_abastecimentos", 0))

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def gerar_pdf(dados: dict) -> bytes:
    try:
        from weasyprint import HTML
    except ImportError:
        raise RuntimeError(
            "WeasyPrint não instalado ou dependências do sistema ausentes. "
            "Consulte o README para instruções de instalação."
        )

    grafico_b64 = _gerar_grafico_barras(dados.get("maquinas", []))

    status_cores = {
        "Econômico": "#22c55e",
        "Normal": "#f59e0b",
        "Alto": "#ef4444",
        "sem_dados": "#94a3b8",
    }

    linhas_tabela = ""
    for m in dados.get("maquinas", []):
        cor = status_cores.get(m.get("status", "sem_dados"), "#94a3b8")
        media_str = f"{m['media']:.2f}" if m.get("media") is not None else "-"
        linhas_tabela += f"""
        <tr>
            <td>{m['nome']}</td>
            <td>{m['tipo']}</td>
            <td>{m['horas']:.1f}</td>
            <td>{m['litros']:.1f}</td>
            <td>{media_str}</td>
            <td><span class="badge" style="background:{cor}">{m.get('status','–')}</span></td>
            <td>{m.get('qtd_abastecimentos', 0)}</td>
        </tr>"""

    html_content = f"""
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 12px; color: #1f2937; margin: 30px; }}
        h1 {{ color: #1e3a5f; font-size: 20px; margin-bottom: 4px; }}
        .periodo {{ color: #6b7280; font-size: 11px; margin-bottom: 20px; }}
        .kpis {{ display: flex; gap: 16px; margin-bottom: 24px; }}
        .kpi {{ background: #f1f5f9; border-left: 4px solid #1e3a5f; padding: 10px 16px; border-radius: 4px; min-width: 120px; }}
        .kpi-label {{ font-size: 10px; color: #6b7280; text-transform: uppercase; }}
        .kpi-value {{ font-size: 18px; font-weight: bold; color: #1e3a5f; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
        th {{ background: #1e3a5f; color: white; padding: 8px 10px; text-align: left; font-size: 11px; }}
        td {{ padding: 7px 10px; border-bottom: 1px solid #e5e7eb; font-size: 11px; }}
        tr:nth-child(even) {{ background: #f9fafb; }}
        .badge {{ color: white; padding: 2px 10px; border-radius: 20px; font-size: 10px; font-weight: bold; }}
        img {{ max-width: 100%; margin-top: 10px; }}
        h2 {{ color: #1e3a5f; font-size: 14px; margin-top: 24px; margin-bottom: 8px; }}
    </style>
    </head>
    <body>
        <h1>Relatório de Gestão de Frota</h1>
        <p class="periodo">Período: {dados.get('periodo', '')}</p>

        <div class="kpis">
            <div class="kpi">
                <div class="kpi-label">Máquinas</div>
                <div class="kpi-value">{dados.get('total_maquinas', 0)}</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Total Horas</div>
                <div class="kpi-value">{dados.get('total_horas', 0):.1f}h</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Total Litros</div>
                <div class="kpi-value">{dados.get('total_litros', 0):.1f}L</div>
            </div>
            <div class="kpi">
                <div class="kpi-label">Média Geral</div>
                <div class="kpi-value">{dados.get('media_geral', 0):.2f} L/h</div>
            </div>
        </div>

        <h2>Consumo por Máquina</h2>
        <table>
            <thead>
                <tr>
                    <th>Máquina</th><th>Tipo</th><th>Horas (h)</th>
                    <th>Litros (L)</th><th>Média (L/h)</th>
                    <th>Status</th><th>Abastecimentos</th>
                </tr>
            </thead>
            <tbody>{linhas_tabela}</tbody>
        </table>

        <h2>Gráfico de Consumo</h2>
        <img src="data:image/png;base64,{grafico_b64}" alt="Gráfico de consumo">
    </body>
    </html>
    """

    return HTML(string=html_content).write_pdf()
